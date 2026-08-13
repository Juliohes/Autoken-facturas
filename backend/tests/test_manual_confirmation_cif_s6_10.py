"""Comportamientos backend de S6.10: veredicto de borrador y excepción de CIF propio.

Se ejercen por HTTP contra Postgres real, sin depender de detalles internos de implementación.
"""

from __future__ import annotations

import io

import httpx
import openpyxl

from tests._invoicing import (
    COUNTERPARTY_CIF,
    INVALID_CIF,
    auth,
    confirm_body,
    confirm_url,
    count_invoices,
    draft_counterparty_verdict_url,
    fetch_invoice,
    seed_confirmable,
    seed_invoice,
)
from tests._reporting import seed_admin_with_company as seed_admin

Api = tuple[httpx.AsyncClient, dict[str, str]]
PANEL_URL = "/api/v1/reporting/invoices"
EXPORT_URL = "/api/v1/reporting/invoices/export"


async def test_c6_veredicto_de_borrador_reemplaza_el_cif_ocr_invalido(authapi: Api) -> None:
    """C6: validar una corrección devuelve el veredicto de los valores enviados, no el del OCR."""
    client, dsns = authapi
    seeded = await seed_confirmable(dsns, client, counterparty_cif=INVALID_CIF, seed_master=False)

    response = await client.post(
        draft_counterparty_verdict_url(seeded["file_id"]),
        headers=auth(seeded["token"]),
        json={"counterparty_tax_id": COUNTERPARTY_CIF, "counterparty_name": "Proveedor SA"},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["counterparty_verdict"]["status"] == "unverified"
    assert body["blocking_reasons"] == []


async def test_c8_confirmacion_revalida_y_devuelve_el_veredicto_estructurado(
    authapi: Api,
) -> None:
    """C8: un body inválido no crea factura y devuelve el resultado actual, estructurado."""
    client, dsns = authapi
    seeded = await seed_confirmable(dsns, client)

    response = await client.post(
        confirm_url(seeded["file_id"]),
        headers=auth(seeded["token"]),
        json=confirm_body(counterparty_cif=INVALID_CIF),
    )

    assert response.status_code == 422, response.text
    assert response.json()["detail"] == {
        "code": "counterparty_cif_invalid",
        "counterparty_verdict": {"status": "invalid", "name_match": None, "official_name": None},
        "blocking_reasons": ["counterparty_cif_invalid"],
    }
    assert await count_invoices(dsns, file_id=seeded["file_id"]) == 0


async def test_c10_user_confirma_excepcion_y_persiste_alerta(authapi: Api) -> None:
    """C10: el user acepta expresamente la ausencia y queda la alerta inmutable para revisar."""
    client, dsns = authapi
    seeded = await seed_confirmable(dsns, client, own_present=False)

    response = await client.post(
        confirm_url(seeded["file_id"]),
        headers=auth(seeded["token"]),
        json={**confirm_body(), "own_tax_id_exception_accepted": True},
    )

    assert response.status_code == 201, response.text
    invoice = await fetch_invoice(dsns, file_id=seeded["file_id"])
    assert invoice is not None
    assert invoice["own_tax_id_missing"] is True
    assert invoice["own_tax_id_exception_confirmed"] is True


async def test_c11_user_sin_casilla_recibe_error_estructurado(authapi: Api) -> None:
    """C11: la excepción no se activa por omisión ni deja una factura a medias."""
    client, dsns = authapi
    seeded = await seed_confirmable(dsns, client, own_present=False)

    response = await client.post(
        confirm_url(seeded["file_id"]), headers=auth(seeded["token"]), json=confirm_body()
    )

    assert response.status_code == 422, response.text
    assert response.json()["detail"] == {
        "code": "own_tax_id_missing",
        "blocking_reasons": ["own_tax_id_missing"],
    }
    assert await count_invoices(dsns, file_id=seeded["file_id"]) == 0


async def test_c12_alerta_se_lista_filtra_y_exporta_solo_en_el_tenant(authapi: Api) -> None:
    """C12: la alerta llega al panel y Excel, y el filtro no cruza el tenant por RLS."""
    client, dsns = authapi
    tenant_id, _admin_id, company_id, token = await seed_admin(dsns, client)
    alert_id = await seed_invoice(
        dsns, tenant_id=tenant_id, company_id=company_id, own_tax_id_missing=True
    )
    await seed_invoice(dsns, tenant_id=tenant_id, company_id=company_id)

    panel = await client.get(PANEL_URL, headers=auth(token, "ilex.localhost"))
    assert panel.status_code == 200, panel.text
    rows = {row["id"]: row for row in panel.json()["items"]}
    assert rows[alert_id]["own_tax_id_missing"] is True

    filtered = await client.get(
        PANEL_URL,
        params={"own_tax_id_missing": "true"},
        headers=auth(token, "ilex.localhost"),
    )
    assert {row["id"] for row in filtered.json()["items"]} == {alert_id}

    exported = await client.get(
        EXPORT_URL,
        params={"own_tax_id_missing": "true"},
        headers=auth(token, "ilex.localhost"),
    )
    assert exported.status_code == 200, exported.text
    sheet = openpyxl.load_workbook(io.BytesIO(exported.content)).active
    assert next(sheet.iter_rows(values_only=True))[-1] == "Revisar CIF propio"
    assert list(sheet.iter_rows(values_only=True))[1][-1] == "Sí"
